# -*- coding: utf-8 -*-
# @Author: ChenYuan
# @Date  : 2020/4/30 11:05
from config.get_path import *
from openpyxl import load_workbook


class GetExcelData():
    def getData(excel, sheet):
        wb = load_workbook(excel)
        ws = wb[sheet]

        # 定义一个列表用来接收所有的测试用例
        all_data = []
        for i in range(2, ws.max_row+1):  # 遍历所有的行
            datas = {}
            datas['case_id'] = ws.cell(i, 1).value  # 获取第I行第一列的第一个格子的值
            datas['title'] = ws.cell(i,2).value
            datas['url'] = ws.cell(i,3).value
            datas['method'] = ws.cell(i,4).value
            datas['headers'] = ws.cell(i,5).value
            datas['data'] = ws.cell(i,6).value
            all_data.append(datas)
        return all_data
